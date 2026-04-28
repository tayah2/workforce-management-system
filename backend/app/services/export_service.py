from io import BytesIO
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Header style constants
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E79')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


def _auto_fit_columns(ws) -> None:
    """Set column widths based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _style_header_row(ws) -> None:
    """Apply bold white text on dark blue background to the first row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def export_payroll_to_excel(payroll_data: List[dict]) -> BytesIO:
    """
    Build an Excel workbook from payroll data and return it as a BytesIO stream.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll Report'

    headers = [
        'Employee ID',
        'Username',
        'Full Name',
        'Hourly Rate (GBP)',
        'Shifts',
        'Total Hours',
        'Regular Hours',
        'Overtime Hours',
        'Regular Pay (GBP)',
        'Overtime Pay (GBP)',
        'Total Pay (GBP)',
    ]
    ws.append(headers)
    _style_header_row(ws)

    for row in payroll_data:
        ws.append([
            row.get('user_id', ''),
            row.get('username', ''),
            row.get('full_name', ''),
            row.get('hourly_rate', 0),
            row.get('shift_count', 0),
            row.get('total_hours', 0),
            row.get('regular_hours', 0),
            row.get('overtime_hours', 0),
            row.get('regular_pay', 0),
            row.get('overtime_pay', 0),
            row.get('total_pay', 0),
        ])

    # Format numeric columns as currency / number
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row[3:]:  # columns D onward are numeric
            cell.number_format = '#,##0.00'

    _auto_fit_columns(ws)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_attendance_to_excel(attendance_data: List[dict]) -> BytesIO:
    """
    Build an Excel workbook from attendance/check-in data and return as BytesIO.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    headers = [
        'Record ID',
        'Employee ID',
        'Username',
        'Full Name',
        'Location',
        'Check-In Time',
        'Check-Out Time',
        'Hours Worked',
        'Break (mins)',
        'Check-In GPS Verified',
        'Check-Out GPS Verified',
        'Notes',
    ]
    ws.append(headers)
    _style_header_row(ws)

    for row in attendance_data:
        # Resolve nested objects if present
        user_info = row.get('user', {}) or {}
        location_info = row.get('location', {}) or {}

        ws.append([
            row.get('id', ''),
            row.get('user_id', user_info.get('id', '')),
            user_info.get('username', ''),
            user_info.get('full_name', ''),
            location_info.get('name', row.get('location_name', '')),
            row.get('check_in_time', ''),
            row.get('check_out_time', ''),
            row.get('hours_worked', 0),
            row.get('break_minutes', 0),
            'Yes' if row.get('is_check_in_verified') else 'No',
            'Yes' if row.get('is_check_out_verified') else 'No',
            row.get('notes', ''),
        ])

    _auto_fit_columns(ws)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
